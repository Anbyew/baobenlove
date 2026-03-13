"""
Wedding Save the Date Email Sender — Gmail API / OAuth2
Reads testList.xlsx, fills in template, and sends emails from bellabenbao@gmail.com

Setup (one-time):
  1. Go to console.cloud.google.com (sign in as bellabenbao@gmail.com)
  2. Create a project → APIs & Services → Library → enable "Gmail API"
  3. APIs & Services → OAuth consent screen → External → fill app name → save
  4. APIs & Services → Credentials → Create Credentials → OAuth client ID
     → Desktop app → Download JSON → save as  emails/credentials.json
  5. Run this script — a browser window will open to authorize
     (token saved to emails/token.json for future runs)

Install deps:
  pip install openpyxl google-auth-oauthlib google-auth-httplib2 google-api-python-client
"""

import base64
import re
import openpyxl
import sys
import webbrowser
import tempfile
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from urllib.parse import urlencode

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# ── CONFIG ────────────────────────────────────────────────────────────────────
SENDER_EMAIL    = "bellabenbao@gmail.com"
BASE_DIR        = Path(__file__).parent.parent
XLSX_PATH       = BASE_DIR / "emails" / "testList.xlsx"
ATTACHMENT_PATH = BASE_DIR / "assets" / "saveTheDate" / "SaveThe Date.png"
CREDS_PATH      = BASE_DIR / "emails" / "credentials.json"
TOKEN_PATH      = BASE_DIR / "emails" / "token.json"

DRY_RUN = True  # ← set to False to actually send
# ─────────────────────────────────────────────────────────────────────────────

SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
EMAIL_SUBJECT = "Save the Date \u2014 October 3, 2026"
WEBSITE_BASE_URL = "https://www.baoben.love/"
INVITE_SOURCE = "save_the_date"


def build_guest_id(name: str, row_number: int) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return f"{normalized or 'guest'}_{row_number}"


def build_email_slug(email: str) -> str:
    local_part = email.split("@", 1)[0]
    normalized = re.sub(r"[^a-z0-9]+", "_", local_part.lower()).strip("_")
    return normalized or "email"


def build_tracked_website_url(guest_id: str) -> str:
    query = urlencode(
        {
            "src": INVITE_SOURCE,
            "guest": guest_id,
            "utm_source": "save_the_date_email",
            "utm_medium": "email",
            "utm_campaign": "save_the_date",
        }
    )
    return f"{WEBSITE_BASE_URL}?{query}"


def get_gmail_service():
    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_PATH), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_PATH.write_text(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def build_html_body(name: str, website_url: str) -> str:
    return f"""
<html><body style="font-family: Georgia, serif; font-size: 16px; line-height: 1.7; color: #2a2a2a; max-width: 640px; margin: auto; padding: 24px;">

<p>Dear {name},</p>

<p>Over the years, our paths have wandered through proofs and algorithms, through long conversations and the quiet pursuit of elegant ideas. Somewhere along the way, we arrived at a realization both simple and profound: <strong>Life is loveliest when its path is walked together.</strong></p>

<p>With great joy, we are getting married.</p>

<p>Our wedding will take place on <strong>October 3, 2026</strong>, at <strong>Longwood Gardens in Kennett Square, Pennsylvania</strong>, and it would mean so much to us to share this celebration with you. ✨</p>

<p>Please find our <strong>Save the Date</strong> for the U.S. celebration attached. A formal invitation with further details will follow in the months ahead.</p>

<p>As our family and friends are scattered across many corners of the world, we will be marking this new chapter with <strong>two celebrations</strong>. You are warmly invited to both, and we would be delighted to celebrate together wherever you may choose to join us.</p>

<p>In the meantime, you may find updates and information on our wedding website:<br>
<strong>Website:</strong> <a href="{website_url}">www.baoben.love</a><br>
<strong>Password:</strong> BellaBenBao2026</p>

<p>As in many beautiful things, elegance and harmony are found not only in ideas, but in the rare joy of gathering those we cherish and celebrating together ❤️</p>

<p>With love,<br>
Yuwei &amp; Ben</p>

<hr style="border: none; border-top: 1px dashed #ccc; margin: 24px 0;">

<p style="font-size: 13px; color: #666;">Additional details regarding the <strong>China celebration</strong> will be shared soon. You might find the following information on tourist visas to China and visa-waiver policies helpful: <a href="https://consular.mfa.gov.cn/VISA/">China Visa Application</a>, <a href="https://www.visaforchina.cn/SYD3_EN/tongzhigonggao/265975107544027136.html">Visa-Free Policy Notice</a></p>
</body></html>
"""


def build_plain_body(name: str, website_url: str) -> str:
    return f"""Dear {name},

Over the years, our paths have wandered through proofs and algorithms, through long conversations and the quiet pursuit of elegant ideas. Somewhere along the way, we arrived at a realization both simple and profound: life is loveliest when its path is walked together.

With great joy, we are getting married.

Our wedding will take place on October 3, 2026, at Longwood Gardens in Kennett Square, Pennsylvania, and it would mean a great deal to us to celebrate this occasion with you.

Please find our Save the Date for the U.S. celebration attached. A formal invitation with further details will follow in the months ahead.

As our family and friends are scattered across many corners of the world, we will be marking this new chapter with two celebrations. You are warmly invited to both, and we would be delighted to celebrate together wherever you may choose to join us.

In the meantime, you may find updates and information on our wedding website:
Website: {website_url}
Password: BellaBenBao2026

As in many beautiful things, elegance and harmony are found not only in ideas, but in the rare joy of gathering those we cherish and celebrating together.

With love,
Yuwei & Ben

---
Additional details regarding the China celebration will be shared soon. You might find the following information on tourist visas to China and visa-waiver policies helpful:
- China Visa Application: https://consular.mfa.gov.cn/VISA/
- Visa-Free Policy Notice: https://www.visaforchina.cn/SYD3_EN/tongzhigonggao/265975107544027136.html
"""


def load_recipients():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    name_col  = headers.index("Name on Envelope")
    email_col = headers.index("Email")

    recipients = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name       = row[name_col]
        emails_raw = row[email_col]
        if not name or not emails_raw:
            continue
        emails = [e.strip() for e in str(emails_raw).split(",") if e.strip()]
        base_guest_id = build_guest_id(str(name), row_number)
        for email in emails:
            guest_id = f"{base_guest_id}_{build_email_slug(email)}"
            recipients.append(
                {
                    "name": name,
                    "email": email,
                    "guest_id": guest_id,
                    "website_url": build_tracked_website_url(guest_id),
                }
            )
    return recipients


def build_message(name, email, website_url):
    msg = MIMEMultipart("mixed")
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = email

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(build_plain_body(name, website_url), "plain"))
    alt.attach(MIMEText(build_html_body(name, website_url), "html"))
    msg.attach(alt)

    with open(ATTACHMENT_PATH, "rb") as f:
        img = MIMEImage(f.read(), name=ATTACHMENT_PATH.name)
        img.add_header("Content-Disposition", "attachment", filename=ATTACHMENT_PATH.name)
        msg.attach(img)

    return msg


def preview():
    """Open a browser preview of the email for the first recipient."""
    recipients = load_recipients()
    if not recipients:
        print("No recipients found.")
        return
    r = recipients[0]
    html = build_html_body(r["name"], r["website_url"])
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8") as f:
        f.write(html)
        path = f.name
    print(f"Preview for: {r['name']} <{r['email']}>")
    print(f"Tracked website: {r['website_url']}")
    webbrowser.open(f"file://{path}")


def main():
    if "--preview" in sys.argv:
        preview()
        return

    recipients = load_recipients()
    print(f"Found {len(recipients)} recipients.\n")

    if DRY_RUN:
        for r in recipients:
            print(f"[DRY RUN] To: {r['email']}  |  Name: {r['name']}  |  Link: {r['website_url']}")
        print("\nDry run complete. Set DRY_RUN = False to actually send.")
        return

    if not CREDS_PATH.exists():
        raise FileNotFoundError(
            f"credentials.json not found at {CREDS_PATH}\n"
            "Follow the setup instructions at the top of this file."
        )

    service = get_gmail_service()

    for r in recipients:
        msg = build_message(r["name"], r["email"], r["website_url"])
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId="me", body={"raw": raw}).execute()
        print(f"Sent to: {r['email']}  |  Name: {r['name']}")

    print("\nAll emails sent!")


if __name__ == "__main__":
    main()
